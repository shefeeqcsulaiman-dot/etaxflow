import os
import re
import base64
import json
import time
from pathlib import Path
from io import BytesIO

from openai import OpenAI

import pdfplumber
from pdf2image import convert_from_path
from PIL import Image
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# ★  CONFIGURATION  — edit these
# ─────────────────────────────────────────────────────────────────────────────
OPENAI_API_KEY = "YOUR_OPENAI_API_KEY"            # ← paste your sk-... key here
INVOICE_FOLDER = "image_invoices"    # folder containing your invoice files
OUTPUT_FILE    = "invoices_output_09_14.xlsx"
BATCH_SIZE     = 10            # files per batch

# Model choice:
#   "gpt-4o-mini"  → cheapest, fast, very accurate  (~$0.001/invoice)
#   "gpt-4o"       → most accurate                   (~$0.01/invoice)
MODEL = "gpt-4o-mini"

OCR_DPI      = 200   # DPI for scanned PDF rendering
MAX_RETRIES  = 3     # retries on network/rate errors
RETRY_WAIT   = 30    # seconds to wait on rate limit (429)
DELAY        = 1     # seconds between files

IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".bmp", ".webp", ".tiff", ".tif"}

# ─────────────────────────────────────────────────────────────────────────────
# EXCEL COLUMNS
# ─────────────────────────────────────────────────────────────────────────────
HEADERS = [
    "Filename",
    "Invoice Date",
    "Invoice Number",
    "Supplier",
    "TRN / VAT #",
    "Bill To",
    "Subtotal (excl. VAT)",
    "Total Discount",
    "VAT Amount",
    "Total Payable",
    "Item Description",
    "Qty",
    "Unit Price",
    "Discount %",
    "Discount Amount",
    "Line Total",
]
COL_WIDTHS = [35, 14, 22, 32, 22, 28, 20, 16, 14, 16, 52, 8, 14, 12, 16, 18]

# ─────────────────────────────────────────────────────────────────────────────
# EXTRACTION PROMPT
# ─────────────────────────────────────────────────────────────────────────────
PROMPT = """You are an invoice data extraction expert.

Read this invoice carefully and extract all data.
Return ONLY a valid JSON object — no markdown, no code fences, no extra text.

{
  "invoice_date":      "date as written on invoice, empty string if not found",
  "invoice_number":    "invoice or tax invoice number, empty string if not found",
  "supplier":          "seller or supplier company name, empty string if not found",
  "trn_vat":           "TRN or VAT registration number of supplier, empty string if not found",
  "bill_to":           "buyer or customer name, empty string if not found",
  "currency":          "3-letter currency code e.g. AED USD EUR — detect from invoice",
  "subtotal_excl_vat": "subtotal before VAT / invoice total before discount as plain number, empty string if not found",
  "total_discount":    "total discount amount for the whole invoice as plain number e.g. 11.00, empty string if no discount shown",
  "vat_amount":        "total VAT amount as plain number, empty string if not found",
  "total_payable":     "final total or balance due including VAT as plain number, empty string if not found",
  "line_items": [
    {
      "description":       "full product or service description exactly as written",
      "qty":               "quantity as string",
      "unit_price":        "unit price before discount as plain number, or Free for free items",
      "discount_pct":      "discount percentage as plain number e.g. 13.00, empty string if none",
      "discount_amount":   "discount amount in currency as plain number e.g. 26.03, empty string if none",
      "line_total":        "final line total including VAT as plain number, 0.00 for free items"
    }
  ]
}

Rules:
- Extract ALL line items including free/sample items, delivery, shipping charges
- total_discount = the summary-level discount shown on the invoice (e.g. "Total Discounted Amount: AED 11.00")
- discount_pct = the % column in the line items table (e.g. 13.00 from "Disc % = 13.00")
- discount_amount = the calculated discount money amount per line if shown, else empty string
- If only discount % is shown and not the amount, leave discount_amount as empty string
- If only discount amount is shown and not the %, leave discount_pct as empty string
- Numbers must be plain digits with decimal point — no currency symbols, no commas
- Keep product descriptions exactly as printed on the invoice
- Use empty string for any value not visible on the invoice
- Do not guess or invent values"""

# ─────────────────────────────────────────────────────────────────────────────
# OPENAI CLIENT SETUP
# ─────────────────────────────────────────────────────────────────────────────
def setup_client() -> OpenAI:
    api_key = OPENAI_API_KEY or os.environ.get("OPENAI_API_KEY", "")
    if not api_key:
        raise ValueError(
            "\nNo OpenAI API key found!\n"
            "  Option 1: Set OPENAI_API_KEY = 'sk-...' at the top of this script\n"
            "  Option 2: Run in CMD:  set OPENAI_API_KEY=sk-...\n"
            "  Get a key at: https://platform.openai.com/api-keys\n"
        )
    return OpenAI(api_key=api_key)

# ─────────────────────────────────────────────────────────────────────────────
# FILE → BASE64 IMAGE(S)
# ─────────────────────────────────────────────────────────────────────────────
def pil_to_b64(img: Image.Image) -> str:
    """Convert PIL Image to base64 PNG string."""
    buf = BytesIO()
    img.save(buf, format="PNG")
    return base64.b64encode(buf.getvalue()).decode("utf-8")


def file_to_parts(path: str) -> list:
    """
    Convert a file into a list of content parts for the OpenAI API.
    Each part is either:
      {"type": "image_url", "image_url": {"url": "data:image/png;base64,..."}}
      {"type": "text",      "text": "..."}
    """
    ext = Path(path).suffix.lower()
    parts = []

    if ext in IMAGE_EXTENSIONS:
        # Single image or multi-frame TIFF
        img = Image.open(path)
        frames = []
        try:
            while True:
                frames.append(img.copy().convert("RGB"))
                img.seek(img.tell() + 1)
        except EOFError:
            pass
        for frame in (frames or [img.convert("RGB")]):
            b64 = pil_to_b64(frame)
            parts.append({
                "type": "image_url",
                "image_url": {"url": f"data:image/png;base64,{b64}", "detail": "high"}
            })

    else:  # PDF
        # Try native text extraction first (fast for digital PDFs)
        native = ""
        with pdfplumber.open(path) as pdf:
            for page in pdf.pages:
                native += page.extract_text() or ""

        if len(native.strip()) > 100:
            # Digital PDF — send as plain text
            parts.append({"type": "text", "text": f"Invoice text content:\n\n{native}"})
        else:
            # Scanned PDF — convert each page to image
            pages = convert_from_path(path, dpi=OCR_DPI)
            for pg in pages:
                b64 = pil_to_b64(pg.convert("RGB"))
                parts.append({
                    "type": "image_url",
                    "image_url": {"url": f"data:image/png;base64,{b64}", "detail": "high"}
                })

    return parts

# ─────────────────────────────────────────────────────────────────────────────
# GPT EXTRACTION  (with retry on rate limit / network errors)
# ─────────────────────────────────────────────────────────────────────────────
def extract_invoice(client: OpenAI, file_path: str) -> dict:
    parts = file_to_parts(file_path)
    if not parts:
        raise ValueError("Could not read file content")

    # Add the extraction prompt as the last text part
    parts.append({"type": "text", "text": PROMPT})

    last_error = None
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            response = client.chat.completions.create(
                model=MODEL,
                messages=[{"role": "user", "content": parts}],
                max_tokens=4096,
                temperature=0,    # deterministic output for data extraction
            )
            raw = response.choices[0].message.content.strip()

            # Strip accidental markdown fences
            raw = re.sub(r"^```(?:json)?\s*", "", raw, flags=re.MULTILINE)
            raw = re.sub(r"```\s*$",           "", raw, flags=re.MULTILINE)
            raw = raw.strip()

            return json.loads(raw)

        except json.JSONDecodeError as e:
            raise  # Don't retry JSON errors — bad prompt response

        except Exception as e:
            last_error = e
            err_str = str(e)

            if "429" in err_str or "rate_limit" in err_str.lower():
                wait = RETRY_WAIT
                m = re.search(r"retry after (\d+)", err_str, re.IGNORECASE)
                if m:
                    wait = int(m.group(1)) + 2
                print(f"    [RATE LIMIT] Waiting {wait}s... (attempt {attempt}/{MAX_RETRIES})")
                time.sleep(wait)

            elif "insufficient_quota" in err_str or "billing" in err_str.lower():
                raise RuntimeError(
                    "OpenAI quota exceeded or billing issue.\n"
                    "  Check your balance at: https://platform.openai.com/usage\n"
                    "  Add credits at: https://platform.openai.com/settings/billing"
                ) from e

            elif "401" in err_str or "invalid_api_key" in err_str.lower():
                raise RuntimeError(
                    "Invalid API key.\n"
                    "  Check your key at: https://platform.openai.com/api-keys"
                ) from e

            else:
                if attempt < MAX_RETRIES:
                    print(f"    [RETRY {attempt}/{MAX_RETRIES}] {e}")
                    time.sleep(5)
                else:
                    raise

    raise last_error

# ─────────────────────────────────────────────────────────────────────────────
# FORMAT HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def fmt(val, currency: str = "AED") -> str:
    """Format a number string as 'AED 1,234.56'."""
    if val is None or str(val).strip() == "":
        return ""
    try:
        return f"{currency} {float(str(val).replace(',', '')):,.2f}"
    except (ValueError, TypeError):
        return str(val)


def build_excel_rows(fname: str, data: dict) -> list:
    cur = (data.get("currency") or "AED").strip().upper()

    # Invoice-level totals — only shown on the FIRST line item row
    totals = {
        "Subtotal (excl. VAT)": fmt(data.get("subtotal_excl_vat"), cur),
        "Total Discount":       fmt(data.get("total_discount"), cur),
        "VAT Amount":           fmt(data.get("vat_amount"), cur),
        "Total Payable":        fmt(data.get("total_payable"), cur),
    }
    # Blank version for subsequent rows
    totals_blank = {k: "" for k in totals}

    hdr_base = {
        "Filename":       fname,
        "Invoice Date":   data.get("invoice_date", ""),
        "Invoice Number": data.get("invoice_number", ""),
        "Supplier":       data.get("supplier", ""),
        "TRN / VAT #":    data.get("trn_vat", ""),
        "Bill To":        data.get("bill_to", ""),
    }

    items = data.get("line_items") or []
    if not items:
        return [{**hdr_base, **totals,
                 "Item Description": "No line items extracted",
                 "Qty": "", "Unit Price": "", "Discount %": "",
                 "Discount Amount": "", "Line Total": ""}]

    rows = []
    for i, item in enumerate(items):
        up   = str(item.get("unit_price", "")      or "")
        lt   = str(item.get("line_total", "")       or "")
        dpct = str(item.get("discount_pct", "")     or "")
        damt = str(item.get("discount_amount", "")  or "")

        # Format discount % — append % sign if it's a plain number
        if dpct:
            try:
                dpct = f"{float(dpct):.2f}%"
            except ValueError:
                pass

        # Invoice totals only on the first line item row
        inv_totals = totals if i == 0 else totals_blank

        rows.append({
            **hdr_base,
            **inv_totals,
            "Item Description": item.get("description", ""),
            "Qty":              str(item.get("qty", "") or ""),
            "Unit Price":       up if up.lower() == "free" else fmt(up, cur),
            "Discount %":       dpct,
            "Discount Amount":  fmt(damt, cur) if damt else "",
            "Line Total":       fmt(lt, cur),
        })
    return rows

# ─────────────────────────────────────────────────────────────────────────────
# EXCEL STYLING
# ─────────────────────────────────────────────────────────────────────────────
def _border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def setup_workbook():
    wb = Workbook()

    # ── Sheet 1: Invoices ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Invoices"
    b = _border()
    ws.merge_cells(f"A1:{get_column_letter(len(HEADERS))}1")
    ws["A1"] = "Invoice Data Extract"
    ws["A1"].font      = Font(name="Arial", bold=True, size=14, color="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    hf    = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    hfill = PatternFill("solid", start_color="1F4E79")
    for col, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = hf; c.fill = hfill; c.border = b
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[2].height = 22
    ws.freeze_panes = "A3"

    # ── Sheet 2: Summary (empty — populated at the end) ──────────────────────
    wb.create_sheet("Summary")

    return wb, ws


def write_summary(wb, all_invoice_data: list):
    """
    Populate the Summary sheet with one row per invoice showing totals,
    plus a grand total row at the bottom.
    all_invoice_data is a list of dicts returned by extract_invoice().
    """
    ws = wb["Summary"]
    b  = _border()

    # ── Title ────────────────────────────────────────────────────────────────
    SUM_HEADERS = [
        "No.", "Filename", "Invoice Date", "Invoice Number",
        "Supplier", "TRN / VAT #", "Bill To",
        "Subtotal (excl. VAT)", "Total Discount",
        "VAT Amount", "Total Payable",
    ]
    SUM_WIDTHS = [5, 35, 14, 22, 32, 22, 26, 20, 16, 14, 16]

    ws.merge_cells(f"A1:{get_column_letter(len(SUM_HEADERS))}1")
    ws["A1"] = "Invoice Summary"
    ws["A1"].font      = Font(name="Arial", bold=True, size=14, color="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill      = PatternFill("solid", start_color="EBF3FB")
    ws.row_dimensions[1].height = 30

    # ── Column headers ───────────────────────────────────────────────────────
    hf    = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    hfill = PatternFill("solid", start_color="1F4E79")
    for col, (h, w) in enumerate(zip(SUM_HEADERS, SUM_WIDTHS), 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = hf; c.fill = hfill; c.border = b
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[2].height = 24
    ws.freeze_panes = "A3"

    def to_float(s):
        if not s:
            return 0.0
        try:
            return float(re.sub(r"[^\d.]", "", str(s)))
        except:
            return 0.0

    # ── Data rows ────────────────────────────────────────────────────────────
    grand_subtotal = grand_discount = grand_vat = grand_total = 0.0

    for i, inv in enumerate(all_invoice_data):
        row_num = i + 3
        cur  = (inv.get("currency") or "AED").strip().upper()
        sub  = fmt(inv.get("subtotal_excl_vat"), cur)
        disc = fmt(inv.get("total_discount"),    cur)
        vat  = fmt(inv.get("vat_amount"),        cur)
        tot  = fmt(inv.get("total_payable"),     cur)

        grand_subtotal += to_float(inv.get("subtotal_excl_vat"))
        grand_discount += to_float(inv.get("total_discount"))
        grand_vat      += to_float(inv.get("vat_amount"))
        grand_total    += to_float(inv.get("total_payable"))

        bg   = "D6E4F0" if i % 2 == 0 else "FFFFFF"
        font = Font(name="Arial", size=10)
        fill_s = PatternFill("solid", start_color=bg)

        row_vals = [
            i + 1,
            inv.get("_filename", ""),
            inv.get("invoice_date", ""),
            inv.get("invoice_number", ""),
            inv.get("supplier", ""),
            inv.get("trn_vat", ""),
            inv.get("bill_to", ""),
            sub, disc, vat, tot,
        ]
        for col, val in enumerate(row_vals, 1):
            c = ws.cell(row=row_num, column=col, value=val)
            c.font = font; c.fill = fill_s; c.border = b
            c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row_num].height = 17

    # ── Grand total row ──────────────────────────────────────────────────────
    gt_row  = len(all_invoice_data) + 3
    cur_def = "AED"
    gt_vals = [
        "", "GRAND TOTAL", "", f"{len(all_invoice_data)} invoices",
        "", "", "",
        fmt(str(grand_subtotal), cur_def),
        fmt(str(grand_discount), cur_def),
        fmt(str(grand_vat),      cur_def),
        fmt(str(grand_total),    cur_def),
    ]
    gt_fill = PatternFill("solid", start_color="BDD7EE")
    gt_font = Font(name="Arial", bold=True, size=11, color="1F4E79")
    for col, val in enumerate(gt_vals, 1):
        c = ws.cell(row=gt_row, column=col, value=val)
        c.font = gt_font; c.fill = gt_fill
        c.border = Border(
            left=Side(style="medium", color="1F4E79"),
            right=Side(style="medium", color="1F4E79"),
            top=Side(style="medium", color="1F4E79"),
            bottom=Side(style="medium", color="1F4E79"),
        )
        c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[gt_row].height = 22

    # ── Legend ───────────────────────────────────────────────────────────────
    leg_row = gt_row + 2
    legends = [
        ("FFF2CC", "7B3F00", "Item has discount (in Invoices sheet)"),
        ("E2EFDA", "375623", "Free / sample item (in Invoices sheet)"),
    ]
    ws.cell(row=leg_row, column=1, value="Legend:").font = Font(
        name="Arial", bold=True, size=9, color="1F4E79")
    for i, (bg, fg, label) in enumerate(legends):
        c = ws.cell(row=leg_row, column=2 + i * 2, value=label)
        c.fill      = PatternFill("solid", start_color=bg)
        c.font      = Font(name="Arial", size=9, color=fg)
        c.alignment = Alignment(horizontal="center")
        c.border    = b



def write_row(ws, row_num: int, data: dict, shade: bool, is_first_row: bool = False):
    b_std = _border()

    desc     = data.get("Item Description", "")
    up       = data.get("Unit Price", "")
    disc_pct = data.get("Discount %", "")

    # Colour logic
    if up == "Free" or "free delivery" in desc.lower():
        bg   = "E2EFDA"   # green  — free item
        font = Font(name="Arial", size=10, color="375623")
    elif disc_pct and disc_pct not in ("0.00%", ""):
        bg   = "FFF2CC"   # yellow — discounted item
        font = Font(name="Arial", size=10, color="7B3F00")
    else:
        bg   = "D6E4F0" if shade else "FFFFFF"
        font = Font(name="Arial", size=10)

    fill_style = PatternFill("solid", start_color=bg)

    # Heavier top border for first row of each invoice
    s_top = Side(style="medium" if is_first_row else "thin",
                 color="1F4E79"  if is_first_row else "BFBFBF")
    s_std = Side(style="thin", color="BFBFBF")
    border = Border(left=s_std, right=s_std, top=s_top, bottom=s_std)

    for col, key in enumerate(HEADERS, 1):
        c = ws.cell(row=row_num, column=col, value=data.get(key, ""))
        c.font   = font
        c.fill   = fill_style
        c.border = border
        c.alignment = Alignment(
            horizontal="center" if key in ("Qty", "Discount %", "Discount Amount") else "left",
            vertical="center",
            wrap_text=(key == "Item Description")
        )
    ws.row_dimensions[row_num].height = 18


def save_workbook(wb, output_file: str) -> str:
    """Save — if file is open in Excel, save with incremented name instead."""
    save_path = output_file
    for attempt in range(5):
        try:
            wb.save(save_path)
            return save_path
        except PermissionError:
            base, ext = os.path.splitext(output_file)
            save_path = f"{base}_{attempt + 1}{ext}"
            print(f"\n  [WARN] '{output_file}' is open in Excel.")
            print(f"         Saving as '{save_path}' instead.")
            print(f"         Tip: Close Excel before running to overwrite the original.")
    raise PermissionError("Could not save — close the Excel file and try again.")

# ─────────────────────────────────────────────────────────────────────────────
# FILE COLLECTION
# ─────────────────────────────────────────────────────────────────────────────
def collect_files(folder: str) -> list:
    if not os.path.isdir(folder):
        print(f"ERROR: Folder '{folder}' not found.")
        print(f"  Expected at: {os.path.abspath(folder)}")
        return []
    files = []
    for fname in sorted(os.listdir(folder)):
        ext = Path(fname).suffix.lower()
        if ext == ".pdf" or ext in IMAGE_EXTENSIONS:
            files.append(os.path.join(folder, fname))
    pdf_c = sum(1 for f in files if f.lower().endswith(".pdf"))
    print(f"Found {len(files)} file(s) — {pdf_c} PDF(s), {len(files)-pdf_c} image(s)")
    return files

# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────
def process_invoices(folder: str = INVOICE_FOLDER, output_file: str = OUTPUT_FILE):
    files = collect_files(folder)
    if not files:
        return

    print("Connecting to OpenAI...")
    client = setup_client()
    print(f"Model  : {MODEL}")
    cost   = "~$0.001" if "mini" in MODEL else "~$0.01"
    print(f"Cost   : {cost} per invoice\n")

    if os.path.exists(output_file):
        wb       = load_workbook(output_file)
        ws       = wb.active
        next_row = ws.max_row + 1
        # Ensure Summary sheet exists when appending
        if "Summary" not in wb.sheetnames:
            wb.create_sheet("Summary")
        print(f"Appending to '{output_file}' from row {next_row}\n")
    else:
        wb, ws   = setup_workbook()
        next_row = 3

    total_ok = total_err = 0
    shade = True
    all_invoice_data = []   # collected for summary sheet

    for batch_start in range(0, len(files), BATCH_SIZE):
        batch = files[batch_start: batch_start + BATCH_SIZE]
        print(f"── Batch {batch_start // BATCH_SIZE + 1} ({len(batch)} file(s)) ──")

        for file_path in batch:
            fname = os.path.basename(file_path)
            ftype = "IMG" if Path(file_path).suffix.lower() in IMAGE_EXTENSIONS else "PDF"
            print(f"  [{ftype}] {fname}")

            try:
                data = extract_invoice(client, file_path)
                rows = build_excel_rows(fname, data)

                print(f"    Invoice # : {data.get('invoice_number') or 'N/A'}")
                print(f"    Supplier  : {data.get('supplier') or 'N/A'}")
                print(f"    Total     : {data.get('currency','AED')} {data.get('total_payable') or 'N/A'}")
                print(f"    Items     : {len(rows)}")

                for i, row in enumerate(rows):
                    write_row(ws, next_row, row, shade, is_first_row=(i == 0))
                    next_row += 1

                # Store for summary — attach filename for reference
                data["_filename"] = fname
                all_invoice_data.append(data)

                total_ok += 1
                shade = not shade

            except json.JSONDecodeError as e:
                total_err += 1
                print(f"    ERROR (GPT returned invalid JSON): {e}")
            except Exception as e:
                total_err += 1
                print(f"    ERROR: {e}")

            time.sleep(DELAY)

        print()

    # ── Write summary sheet ──────────────────────────────────────────────────
    if all_invoice_data:
        print("Writing Summary sheet...")
        write_summary(wb, all_invoice_data)

    saved_path = save_workbook(wb, output_file)
    print("─" * 55)
    print(f"Done!  {total_ok} invoice(s) → '{saved_path}'")
    print(f"       2 sheets: 'Invoices' (all line items) + 'Summary' (one row per invoice)")
    if total_err:
        print(f"Errors: {total_err} file(s) failed.")


if __name__ == "__main__":
    process_invoices()
